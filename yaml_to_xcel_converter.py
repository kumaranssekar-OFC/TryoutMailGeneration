import os
import copy
import yaml
import openpyxl

TEMPLATE_XLSX = "init.xlsx"          # your existing template with Validlists
YAML_FILE = "tryout.yaml"
OUT_XLSX = "init_from_yaml.xlsx"

# Excel headers in your template (row 1) include a trailing space in Jira_Main_Task
EXCEL_HEADERS = [
    "Part_Numbers",
    "Predecessor_PN",
    "SW_Version",
    "Jira_Main_Task ",   # IMPORTANT: trailing space as in template [1](https://bosch-my.sharepoint.com/personal/euj1cob_bosch_com/_layouts/15/Doc.aspx?sourcedoc=%7B892717A4-2A83-4C8B-8EC0-2E26D054FCB0%7D&file=init.xlsx&action=default&mobileredirect=true)
    "Jira_TO_Task",
    "PD_Version",
    "CD_Version",
    "ProductType",
    "Base_SW",
    "BaseSW_Task",
    "Release_Type",
    "Docushare_CollectionID",
    "Tryout_Location",
    "FCID_Version",
    "HW_List",
    "Private_Key",
    "Jira_Access_Token",
    "Task_Name",
]

# Allow YAML keys without the trailing space
KEY_ALIASES = {
    "Jira_Main_Task": "Jira_Main_Task ",   # map YAML key -> Excel header
}

# Optional: Load secrets from environment if not provided in YAML
ENV_FALLBACKS = {
    "Private_Key": "TRYOUT_PRIVATE_KEY",
    "Jira_Access_Token": "JIRA_ACCESS_TOKEN",
}

def normalize_key(k: str) -> str:
    """Map YAML keys to Excel header names (e.g., handle Jira_Main_Task space)."""
    return KEY_ALIASES.get(k, k)

def load_yaml(yaml_path: str) -> dict:
    with open(yaml_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}

def get_env_if_missing(data: dict, field: str) -> str:
    """Return data[field] if present else ENV fallback if configured."""
    if field in data and data[field] not in (None, ""):
        return data[field]
    env_name = ENV_FALLBACKS.get(field)
    if env_name:
        return os.getenv(env_name, "")
    return ""

def main():
    # Load YAML
    cfg = load_yaml(YAML_FILE)
    defaults = cfg.get("defaults", {}) or {}
    runs = cfg.get("runs", []) or []

    # Load template workbook so we preserve Validlists
    wb = openpyxl.load_workbook(TEMPLATE_XLSX)
    if "init" not in wb.sheetnames:
        raise RuntimeError("Template must contain 'init' sheet.")
    ws = wb["init"]

    # Validate template headers match expected (row 1)
    template_headers = [ws.cell(1, c).value for c in range(1, len(EXCEL_HEADERS) + 1)]
    if template_headers != EXCEL_HEADERS:
        raise RuntimeError(
            "Template headers do not match expected structure.\n"
            f"Found: {template_headers}\nExpected: {EXCEL_HEADERS}"
        )

    # Clear existing data rows (row 2 onwards) for the used columns
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        for c in range(1, len(EXCEL_HEADERS) + 1):
            ws.cell(r, c).value = None

    # Write each run as a row
    out_row = 2
    for run in runs:
        # Merge defaults + run (run overrides defaults)
        merged = copy.deepcopy(defaults)
        merged.update(run or {})

        # Normalize YAML keys to match Excel headers
        normalized = {}
        for k, v in merged.items():
            normalized[normalize_key(k)] = v

        # Fill secrets from ENV if missing (recommended)
        normalized["Private_Key"] = get_env_if_missing(normalized, "Private_Key")
        normalized["Jira_Access_Token"] = get_env_if_missing(normalized, "Jira_Access_Token")

        # Write in header order
        for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
            ws.cell(out_row, col_idx).value = normalized.get(header, None)

        out_row += 1

    # Save output
    wb.save(OUT_XLSX)
    print(f"Generated: {OUT_XLSX}")

if __name__ == "__main__":
    main()